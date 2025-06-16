# -*- coding: utf-8 -*-
from django.shortcuts import render
from django.contrib.auth.decorators import login_required
from django.views.generic import ListView, View
from django.http import HttpResponse
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework.permissions import IsAuthenticated
from rest_framework.decorators import api_view, permission_classes
from django.utils import timezone
from auditlog.models import LogEntry as AuditLog
import csv
import openpyxl
from openpyxl.utils import get_column_letter
from reportlab.pdfgen import canvas
import tempfile

@login_required
def report_dashboard(request):
    # This will be expanded later
    return render(request, 'reports/dashboard.html')

class ReportListView(ListView):
    model = Report
    template_name = 'reports/report_list.html'
    context_object_name = 'reports'
    paginate_by = 20

class ReportExportCSV(View):
    def get(self, request, *args, **kwargs):
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = 'attachment; filename="reports.csv"'
        writer = csv.writer(response)
        writer.writerow(['ID', 'Название', 'Дата', 'Автор'])
        # MOCK: заменить на реальные данные отчетов
        writer.writerow([1, 'Форма 30', timezone.now().date(), request.user.username])
        writer.writerow([2, 'Форма 12', timezone.now().date(), request.user.username])
        AuditLog.objects.log_create(
            actor=request.user,
            action='export_report_csv',
            remote_addr=request.META.get('REMOTE_ADDR')
        )
        return response

# REST API для получения отчета формы 30 (пример)
@api_view(['GET'])
@permission_classes([IsAuthenticated])
def report_30_api(request):
    # MOCK: заменить на реальную агрегацию данных
    data = {
        'organization': 'Медицинский центр',
        'period': str(timezone.now().date()),
        'patients_count': 123,
        'services_count': 456,
    }
    AuditLog.objects.log_create(
        actor=request.user,
        action='view_report_30',
        remote_addr=request.META.get('REMOTE_ADDR')
    )
    return Response(data)

# Выгрузка формы 30 в Excel
@login_required
def report_30_excel(request):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Форма 30'
    ws.append(['Организация', 'Период', 'Пациентов', 'Услуг'])
    ws.append(['Медицинский центр', str(timezone.now().date()), 123, 456])
    for col in range(1, 5):
        ws.column_dimensions[get_column_letter(col)].width = 20
    with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp:
        wb.save(tmp.name)
        tmp.seek(0)
        response = HttpResponse(tmp.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="report_30.xlsx"'
        AuditLog.objects.log_create(
            actor=request.user,
            action='export_report_30_excel',
            remote_addr=request.META.get('REMOTE_ADDR')
        )
        return response

# Выгрузка формы 30 в PDF
@login_required
def report_30_pdf(request):
    with tempfile.NamedTemporaryFile(delete=False, suffix='.pdf') as tmp:
        c = canvas.Canvas(tmp.name)
        c.drawString(100, 800, 'Форма 30')
        c.drawString(100, 780, f'Организация: Медицинский центр')
        c.drawString(100, 760, f'Период: {timezone.now().date()}')
        c.drawString(100, 740, f'Пациентов: 123')
        c.drawString(100, 720, f'Услуг: 456')
        c.save()
        tmp.seek(0)
        response = HttpResponse(tmp.read(), content_type='application/pdf')
        response['Content-Disposition'] = 'attachment; filename="report_30.pdf"'
        AuditLog.objects.log_create(
            actor=request.user,
            action='export_report_30_pdf',
            remote_addr=request.META.get('REMOTE_ADDR')
        )
        return response
